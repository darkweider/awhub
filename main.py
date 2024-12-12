import asyncio
import aiosqlite
import aiofiles
import pickle
import logging
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from telebot.async_telebot import AsyncTeleBot
from telebot import types
from telebot.types import InputMediaPhoto

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("bot.log"),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

API_TOKEN = '7289461911:AAHQCesDv_AWBjA7Nlz5TfSy2XlQ-dhac6Q'
bot = AsyncTeleBot(API_TOKEN)

# Путь к файлам базы данных и Excel
DB_PATH = 'artfind.db'
EXCEL_FILE = 'start.xlsx'
USER_STATES_FILE = 'user_states.pkl'
PHOTOS_DIR = 'photos'

# Используем словарь для хранения состояния каждого пользователя
user_states = {}

# Синхронная функция для работы с Excel
def append_to_excel_sync(file, user_id, date_str, time_str):
    wb = load_workbook(file)
    ws = wb['date']
    ws.append([str(user_id), date_str, time_str])
    wb.save(file)
    wb.close()

# Асинхронная функция для инициализации Excel файла
async def init_excel(file):
    try:
        load_workbook(file)
        logger.info(f"Excel файл '{file}' загружен.")
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = 'date'
        ws.append(['tgid', 'date', 'time'])
        wb.save(file)
        wb.close()
        logger.info(f"Создан новый Excel файл: {file}")

# Асинхронная функция для инициализации базы данных
async def init_db():
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute('''
            CREATE TABLE IF NOT EXISTS artist (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tgid TEXT UNIQUE,
                name TEXT,
                style TEXT,
                username TEXT
            )
        ''')
        await db.execute('''
            CREATE TABLE IF NOT EXISTS customer (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tgid TEXT UNIQUE,
                name TEXT,
                info TEXT,
                username TEXT
            )
        ''')
        await db.execute('''
            CREATE TABLE IF NOT EXISTS ratings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                artist_id TEXT,
                likes INTEGER DEFAULT 0,
                dislikes INTEGER DEFAULT 0,
                FOREIGN KEY (artist_id) REFERENCES artist (tgid)
            )
        ''')
        await db.execute('''
            CREATE TABLE IF NOT EXISTS user_ratings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id TEXT,
                artist_id TEXT,
                action TEXT,
                FOREIGN KEY (customer_id) REFERENCES customer (tgid),
                FOREIGN KEY (artist_id) REFERENCES artist (tgid)
            )
        ''')
        await db.execute('''
            CREATE TABLE IF NOT EXISTS favorites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id TEXT,
                artist_id TEXT,
                FOREIGN KEY (customer_id) REFERENCES customer (tgid),
                FOREIGN KEY (artist_id) REFERENCES artist (tgid)
            )
        ''')
        await db.commit()
        logger.info("База данных инициализирована.")

# Асинхронная функция для загрузки состояний пользователей из файла
def load_user_states_sync():
    try:
        with open(USER_STATES_FILE, 'rb') as f:
            data = pickle.load(f)
        logger.info("Состояния пользователей успешно загружены из файла.")
        return data
    except (FileNotFoundError, pickle.UnpicklingError) as e:
        logger.warning(f"Не удалось загрузить user_states: {e}")
        return {}

async def load_user_states():
    global user_states
    loaded_data = await asyncio.to_thread(load_user_states_sync)
    if isinstance(loaded_data, dict):
        user_states.update(loaded_data)
    logger.info(f"Текущее состояние user_states: {user_states}")

# Асинхронная функция для сохранения состояний пользователей в файл
def save_user_states_sync():
    try:
        with open(USER_STATES_FILE, 'wb') as f:
            pickle.dump(user_states, f)
        logger.info("user_states успешно сохранены.")
    except Exception as e:
        logger.error(f"Не удалось сохранить user_states: {e}")

async def save_user_states_async():
    await asyncio.to_thread(save_user_states_sync)

# Инициализируем Excel файл, базу данных и загружаем состояния пользователей
async def initialize():
    await init_excel(EXCEL_FILE)
    await init_db()
    await load_user_states()

# Запуск инициализации
asyncio.run(initialize())

@bot.message_handler(commands=['start'])
async def send_welcome(message):
    user_id = message.from_user.id
    logger.info(f"Пользователь {user_id} начал взаимодействие.")
    current_time = datetime.now()
    date_str = current_time.strftime("%d-%m-%Y")
    time_str = current_time.strftime("%H:%M:%S")

    # Работа с Excel файлом в отдельном потоке
    await asyncio.to_thread(append_to_excel_sync, EXCEL_FILE, user_id, date_str, time_str)
    logger.info(f"Запись пользователя {user_id} в Excel файл.")

    # Работа с базой данных
    async with aiosqlite.connect(DB_PATH) as db:
        # Проверка, является ли пользователь художником или заказчиком
        async with db.execute("SELECT 1 FROM artist WHERE tgid = ?", (str(user_id),)) as cursor:
            ta = await cursor.fetchone()
        async with db.execute("SELECT 1 FROM customer WHERE tgid = ?", (str(user_id),)) as cursor:
            tc = await cursor.fetchone()

        if not ta and not tc:
            # Новый пользователь
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            art = types.KeyboardButton("Художник")
            cus = types.KeyboardButton("Заказчик")
            markup.row(art, cus)
            await bot.send_message(
                message.chat.id,
                f"Здравствуйте, {message.from_user.first_name}! Давайте определимся, кто вы?",
                reply_markup=markup
            )
            # Сохраняем состояние пользователя
            user_states[user_id] = {'role': None, 'awaiting': 'choose_role'}
            await save_user_states_async()
            logger.info(f"Пользователь {user_id} новый, ожидает выбора роли.")
        elif tc:
            # Возвращение заказчика
            async with db.execute('SELECT name, info FROM customer WHERE tgid = ?', (str(user_id),)) as cursor:
                res = await cursor.fetchone()
            if res:
                name, info = res
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                prof = types.KeyboardButton("Мой профиль")
                artsc = types.KeyboardButton("Найти художника")
                fav = types.KeyboardButton("Избранное")
                markup.row(prof, artsc)
                markup.row(fav)
                await bot.send_message(
                    message.chat.id,
                    f"Здравствуйте, {name}!",
                    reply_markup=markup
                )
                # Обновляем состояние
                user_states[user_id] = {'role': 'customer', 'awaiting': 'main_menu'}
                await save_user_states_async()
                logger.info(f"Пользователь {user_id} вернулся как заказчик.")
        elif ta:
            # Возвращение художника
            async with db.execute('SELECT name, style FROM artist WHERE tgid = ?', (str(user_id),)) as cursor:
                res = await cursor.fetchone()
            if res:
                name, style = res
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                prof = types.KeyboardButton("Мой профиль")
                markup.row(prof)
                await bot.send_message(
                    message.chat.id,
                    f"Здравствуйте, {name}!",
                    reply_markup=markup
                )
                # Обновляем состояние
                user_states[user_id] = {'role': 'artist', 'awaiting': 'main_menu'}
                await save_user_states_async()
                logger.info(f"Пользователь {user_id} вернулся как художник.")
        else:
            # Неизвестный случай
            await bot.send_message(message.chat.id, "Произошла ошибка. Пожалуйста, попробуйте позже.")
            logger.error(f"Пользователь {user_id} находится в неизвестном состоянии.")

    logger.info(f"Текущее состояние пользователя {user_id}: {user_states.get(user_id)}")

@bot.message_handler(func=lambda message: True)
async def handle_all_messages(message):
    user_id = message.from_user.id
    state = user_states.get(user_id, {})
    role = state.get('role')
    awaiting = state.get('awaiting')

    logger.info(f"Обработка сообщения от пользователя {user_id}: {message.text}. Состояние: {state}")

    if awaiting == 'choose_role':
        if message.text.lower() == "художник":
            user_states[user_id] = {'role': 'artist', 'awaiting': 'artist_name'}
            await save_user_states_async()
            await ask_artist_details(message)
            logger.info(f"Пользователь {user_id} выбрал роль 'Художник'.")
        elif message.text.lower() == "заказчик":
            user_states[user_id] = {'role': 'customer', 'awaiting': 'customer_name'}
            await save_user_states_async()
            await ask_customer_details(message)
            logger.info(f"Пользователь {user_id} выбрал роль 'Заказчик'.")
        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            art = types.KeyboardButton("Художник")
            cus = types.KeyboardButton("Заказчик")
            markup.row(art, cus)
            await bot.send_message(
                message.chat.id,
                "Пожалуйста, выберите одну из опций ниже.",
                reply_markup=markup
            )
            logger.warning(f"Пользователь {user_id} выбрал неверную опцию роли.")
    elif awaiting == 'main_menu' and role == 'customer':
        if message.text.lower() == "мой профиль":
            await show_customer_profile(message, state)
            logger.info(f"Пользователь {user_id} запросил свой профиль.")
        elif message.text.lower() == "найти художника":
            await search_artists(message)
            logger.info(f"Пользователь {user_id} выбрал поиск художника.")
        elif message.text.lower() == "избранное":
            await show_favorites(message)
            logger.info(f"Пользователь {user_id} выбрал просмотр избранного.")
        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            prof = types.KeyboardButton("Мой профиль")
            artsc = types.KeyboardButton("Найти художника")
            fav = types.KeyboardButton("Избранное")
            markup.row(prof, artsc)
            markup.row(fav)
            await bot.send_message(
                message.chat.id,
                "Пожалуйста, выберите одну из опций ниже.",
                reply_markup=markup
            )
            logger.warning(f"Пользователь {user_id} выбрал неверную опцию в главном меню заказчика.")
    elif awaiting == 'main_menu' and role == 'artist':
        if message.text.lower() == "мой профиль":
            await show_artist_profile(message, state)
            logger.info(f"Пользователь {user_id} запросил свой профиль.")
        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            prof = types.KeyboardButton("Мой профиль")
            markup.row(prof)
            await bot.send_message(
                message.chat.id,
                "Пожалуйста, выберите одну из опций ниже.",
                reply_markup=markup
            )
            logger.warning(f"Пользователь {user_id} выбрал неверную опцию в главном меню художника.")
    elif awaiting in ['artist_name', 'artist_info', 'artist_photo', 'artist_save_confirmation', 'go_out', ]:
        await handle_artist_messages(message, state)
    elif awaiting in ['customer_name', 'customer_info', 'customer_photo', 'customer_save_confirmation', 'go_out']:
        await handle_customer_messages(message, state)
    elif awaiting == 'go_out':
        if message.text.lower() == "изменить профиль":
            if role == 'customer':
                user_states[user_id] = {'role': 'customer', 'awaiting': 'customer_name'}
                await save_user_states_async()
                await ask_customer_details(message)
                logger.info(f"Пользователь {user_id} выбрал изменение профиля заказчика.")
            elif role == 'artist':
                user_states[user_id] = {'role': 'artist', 'awaiting': 'artist_name'}
                await save_user_states_async()
                await ask_artist_details(message)
                logger.info(f"Пользователь {user_id} выбрал изменение профиля художника.")
        elif message.text.lower() == "удалить профиль":
            if role == 'customer':
                await delete_customer_profile(message, state)
                logger.info(f"Пользователь {user_id} выбрал удаление профиля заказчика.")
            elif role == 'artist':
                await delete_artist_profile(message, state)
                logger.info(f"Пользователь {user_id} выбрал удаление профиля художника.")
        elif message.text.lower() == "на главную":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} вернулся на главную из состояния 'go_out'.")
        else:
            await bot.send_message(message.chat.id, "Пожалуйста, выберите одну из опций ниже.")
            logger.warning(f"Пользователь {user_id} выбрал неверную опцию в состоянии 'go_out'.")
    else:
        await bot.send_message(message.chat.id, "Произошла ошибка состояния. Пожалуйста, начните заново командой /start.")
        logger.error(f"Пользователь {user_id} находится в неизвестном состоянии: {state}")
        user_states.pop(user_id, None)
        await save_user_states_async()

async def ask_customer_details(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    delit = types.KeyboardButton("На главную")
    markup.add(delit)
    await bot.send_message(message.chat.id, "Как могу к вам обращаться?", reply_markup=markup)
    user_states[message.from_user.id]['awaiting'] = 'customer_name'
    await save_user_states_async()
    logger.info(f"Пользователь {message.from_user.id} ожидает ввода имени заказчика.")

async def ask_artist_details(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    delit = types.KeyboardButton("На главную")
    markup.add(delit)
    await bot.send_message(message.chat.id, "Как могу к вам обращаться?", reply_markup=markup)
    user_states[message.from_user.id]['awaiting'] = 'artist_name'
    await save_user_states_async()
    logger.info(f"Пользователь {message.from_user.id} ожидает ввода имени художника.")

async def handle_customer_messages(message, state):
    user_id = message.from_user.id
    awaiting = state.get('awaiting')

    if awaiting == 'customer_name':
        if message.text.lower() == "на главную":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} вернулся на главную из состояния 'customer_name'.")
        else:
            state['name'] = message.text.strip()
            state['username'] = message.from_user.username
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            delit = types.KeyboardButton("На главную")
            markup.add(delit)
            await bot.send_message(
                message.chat.id,
                f"Хорошо, {state['name']}, теперь напишите о себе.",
                reply_markup=markup
            )
            state['awaiting'] = 'customer_info'
            await save_user_states_async()
            logger.info(f"Пользователь {user_id} ввел имя заказчика: {state['name']}")

    elif awaiting == 'customer_info':
        if message.text.lower() == "на главную":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} вернулся на главную из состояния 'customer_info'.")
        else:
            state['info'] = message.text.strip()
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            delit = types.KeyboardButton("На главную")
            markup.add(delit)
            await bot.send_message(
                message.chat.id,
                "Отлично! Теперь скиньте аватарку для вашего профиля.",
                reply_markup=markup
            )
            state['awaiting'] = 'customer_photo'
            state['photos_saved'] = False  # Добавляем флаг для отслеживания сохранения фотографий
            await save_user_states_async()
            logger.info(f"Пользователь {user_id} ввел информацию о себе: {state['info']}")

    elif awaiting == 'customer_photo':
        if message.photo:
            await save_photos(message)
            if not state.get('photos_saved', False):
                state['photos_saved'] = True  # Устанавливаем флаг после сохранения фотографий
                await save_user_states_async()
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                save_btn = types.KeyboardButton("Да")
                cancel_btn = types.KeyboardButton("Нет")
                markup.row(save_btn, cancel_btn)
                await bot.send_message(
                    message.chat.id,
                    "Фотографии сохранены. Сохранить профиль?",
                    reply_markup=markup
                )
                state['awaiting'] = 'customer_save_confirmation'
                await save_user_states_async()
                logger.info(f"Пользователь {user_id} отправил аватарку и ожидает подтверждения сохранения.")
        elif message.text.lower() == "на главную":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} вернулся на главную из состояния 'customer_photo'.")
        else:
            await bot.send_message(message.chat.id, "Пожалуйста, отправьте фотографию или выберите 'На главную'.")
            logger.warning(f"Пользователь {user_id} отправил неверный тип данных в состоянии 'customer_photo'.")

    elif awaiting == 'customer_save_confirmation':
        if message.text.lower() == "да":
            await save_customer_to_db(message, state)
            logger.info(f"Пользователь {user_id} подтвердил сохранение профиля заказчика.")
        elif message.text.lower() == "нет":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} отменил сохранение профиля заказчика.")
        else:
            await bot.send_message(message.chat.id, "Пожалуйста, выберите 'Да' или 'Нет'.")
            logger.warning(f"Пользователь {user_id} отправил неверный ответ в состоянии 'customer_save_confirmation'.")

    elif awaiting == 'go_out':
        if message.text.lower() == "изменить профиль":
            user_states[user_id] = {'role': 'customer', 'awaiting': 'customer_name'}
            await save_user_states_async()
            await ask_customer_details(message)
            logger.info(f"Пользователь {user_id} выбрал изменение профиля заказчика.")
        elif message.text.lower() == "удалить профиль":
            await delete_customer_profile(message, state)
            logger.info(f"Пользователь {user_id} выбрал удаление профиля заказчика.")
        elif message.text.lower() == "на главную":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} вернулся на главную из состояния 'go_out'.")
        else:
            await bot.send_message(message.chat.id, "Пожалуйста, выберите одну из опций ниже.")
            logger.warning(f"Пользователь {user_id} выбрал неверную опцию в состоянии 'go_out'.")
    else:
        await bot.send_message(message.chat.id, "Неизвестное состояние. Пожалуйста, начните заново командой /start.")
        logger.error(f"Пользователь {user_id} находится в неизвестном состоянии: {state}")
        user_states.pop(user_id, None)
        await save_user_states_async()

async def handle_artist_messages(message, state):
    user_id = message.from_user.id
    awaiting = state.get('awaiting')

    if awaiting == 'artist_name':
        if message.text.lower() == "на главную":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} вернулся на главную из состояния 'artist_name'.")
        else:
            state['name'] = message.text.strip()
            state['username'] = message.from_user.username
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            delit = types.KeyboardButton("На главную")
            markup.add(delit)
            await bot.send_message(
                message.chat.id,
                f"Хорошо, {state['name']}, теперь напишите о себе. Добавьте ссылки на ваши профили и хэштеги ",
                reply_markup=markup
            )
            state['awaiting'] = 'artist_info'
            await save_user_states_async()
            logger.info(f"Пользователь {user_id} ввел имя художника: {state['name']}")

    elif awaiting == 'artist_info':
        if message.text.lower() == "на главную":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} вернулся на главную из состояния 'artist_info'.")
        else:
            state['info'] = message.text.strip()
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            delit = types.KeyboardButton("На главную")
            markup.add(delit)
            await bot.send_message(
                message.chat.id,
                "Отлично! Теперь скиньте пример вашей работы.",
                reply_markup=markup
            )
            state['awaiting'] = 'artist_photo'
            state['photos_saved'] = False  # Добавляем флаг для отслеживания сохранения фотографий
            await save_user_states_async()
            logger.info(f"Пользователь {user_id} ввел информацию о себе: {state['info']}")

    elif awaiting == 'artist_photo':
        if message.photo:
            await save_photos(message)
            if not state.get('photos_saved', False):
                state['photos_saved'] = True  # Устанавливаем флаг после сохранения фотографий
                await save_user_states_async()
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                save_btn = types.KeyboardButton("Сохранить")
                cancel_btn = types.KeyboardButton("Удалить")
                markup.row(save_btn, cancel_btn)
                await bot.send_message(
                    message.chat.id,
                    "Фотографии сохранены. Сохранить профиль?",
                    reply_markup=markup
                )
                state['awaiting'] = 'artist_save_confirmation'
                await save_user_states_async()
                logger.info(f"Пользователь {user_id} отправил пример работы и ожидает подтверждения сохранения.")
        elif message.text.lower() == "на главную":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} вернулся на главную из состояния 'artist_photo'.")
        else:
            await bot.send_message(message.chat.id, "Пожалуйста, отправьте фотографию или выберите 'На главную'.")
            logger.warning(f"Пользователь {user_id} отправил неверный тип данных в состоянии 'artist_photo'.")

    elif awaiting == 'artist_save_confirmation':
        if message.text.lower() == "сохранить":
            await save_artist_to_db(message, state)
            logger.info(f"Пользователь {user_id} подтвердил сохранение профиля художника.")
        elif message.text.lower() == "удалить":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} отменил сохранение профиля художника.")
        else:
            await bot.send_message(message.chat.id, "Пожалуйста, выберите 'Сохранить' или 'Удалить'.")
            logger.warning(f"Пользователь {user_id} отправил неверный ответ в состоянии 'artist_save_confirmation'.")
    elif awaiting == 'go_out':
        if message.text.lower() == "изменить профиль":
                user_states[user_id] = {'role': 'artist', 'awaiting': 'artist_name'}
                await save_user_states_async()
                await ask_artist_details(message)
                logger.info(f"Пользователь {user_id} выбрал изменение профиля художника.")
        elif message.text.lower() == "удалить профиль":
                await delete_artist_profile(message, state)
                logger.info(f"Пользователь {user_id} выбрал удаление профиля художника.")
        elif message.text.lower() == "на главную":
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)
            logger.info(f"Пользователь {user_id} вернулся на главную из состояния 'go_out'.")
        else:
            await bot.send_message(message.chat.id, "Пожалуйста, выберите одну из опций ниже.")
            logger.warning(f"Пользователь {user_id} выбрал неверную опцию в состоянии 'go_out'.")
    else:
        await bot.send_message(message.chat.id, "Неизвестное состояние. Пожалуйста, начните заново командой /start.")
        logger.error(f"Пользователь {user_id} находится в неизвестном состоянии: {state}")
        user_states.pop(user_id, None)
        await save_user_states_async()

async def save_customer_to_db(message, state):
    user_id = message.from_user.id
    name = state.get('name')
    info = state.get('info')
    username = state.get('username')

    async with aiosqlite.connect(DB_PATH) as db:
        try:
            await db.execute(
                'INSERT INTO customer (tgid, name, info, username) VALUES (?, ?, ?, ?) ON CONFLICT(tgid) DO UPDATE SET name=excluded.name, info=excluded.info, username=excluded.username',
                (str(user_id), name, info, username)
            )
            await db.commit()
            await bot.send_message(message.chat.id, "Ваш профиль сохранен.", reply_markup=types.ReplyKeyboardRemove())
            logger.info(f"Профиль заказчика пользователя {user_id} сохранен в базе данных.")
        except aiosqlite.IntegrityError:
            await bot.send_message(message.chat.id,
                                   "Произошла ошибка при сохранении. Возможно, профиль уже существует.")
            logger.error(f"Ошибка сохранения профиля заказчика пользователя {user_id}: профиль уже существует.")
        finally:
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)

    # Удаляем старые фотографии пользователя, если директория существует
    user_dir = os.path.join(PHOTOS_DIR, str(user_id))
    if os.path.exists(user_dir):
        for file_name in os.listdir(user_dir):
            file_path = os.path.join(user_dir, file_name)
            os.remove(file_path)
        logger.info(f"Старые фотографии пользователя {user_id} удалены.")
    else:
        os.makedirs(user_dir)
        logger.info(f"Создана директория для фотографий пользователя {user_id}.")

    # Сохраняем новые фотографии
    if 'photos' in state:
        for photo in state['photos']:
            file_name = os.path.join(user_dir, f"{photo['file_id']}.jpg")
            async with aiofiles.open(file_name, 'wb') as f:
                await f.write(photo['file_data'])
        logger.info(f"Новые фотографии пользователя {user_id} сохранены.")

async def save_artist_to_db(message, state):
    user_id = message.from_user.id
    name = state.get('name')
    info = state.get('info')
    username = state.get('username')

    async with aiosqlite.connect(DB_PATH) as db:
        try:
            await db.execute(
                'INSERT INTO artist (tgid, name, style, username) VALUES (?, ?, ?, ?) ON CONFLICT(tgid) DO UPDATE SET name=excluded.name, style=excluded.style, username=excluded.username',
                (str(user_id), name, info, username)
            )
            await db.commit()
            await bot.send_message(message.chat.id, "Ваш профиль сохранен.", reply_markup=types.ReplyKeyboardRemove())
            logger.info(f"Профиль художника пользователя {user_id} сохранен в базе данных.")
        except aiosqlite.IntegrityError:
            await bot.send_message(message.chat.id,
                                   "Произошла ошибка при сохранении. Возможно, профиль уже существует.")
            logger.error(f"Ошибка сохранения профиля художника пользователя {user_id}: профиль уже существует.")
        finally:
            user_states.pop(user_id, None)
            await save_user_states_async()
            await send_welcome(message)

    # Удаляем старые фотографии пользователя, если директория существует
    user_dir = os.path.join(PHOTOS_DIR, str(user_id))
    if os.path.exists(user_dir):
        for file_name in os.listdir(user_dir):
            file_path = os.path.join(user_dir, file_name)
            os.remove(file_path)
        logger.info(f"Старые фотографии пользователя {user_id} удалены.")
    else:
        os.makedirs(user_dir)
        logger.info(f"Создана директория для фотографий пользователя {user_id}.")

    # Сохраняем новые фотографии
    if 'photos' in state:
        for photo in state['photos']:
            file_name = os.path.join(user_dir, f"{photo['file_id']}.jpg")
            async with aiofiles.open(file_name, 'wb') as f:
                await f.write(photo['file_data'])
        logger.info(f"Новые фотографии пользователя {user_id} сохранены.")

async def show_customer_profile(message, state):
    user_id = message.from_user.id
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute('SELECT name, info, username FROM customer WHERE tgid = ?', (str(user_id),)) as cursor:
            res = await cursor.fetchone()
    if res:
        name, info, username = res
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        delete_btn = types.KeyboardButton("Удалить профиль")
        delit = types.KeyboardButton("На главную")
        markup.row(delete_btn)
        markup.row(delit)
        user_dir = os.path.join(PHOTOS_DIR, str(user_id))
        media_group = []

        photos = [os.path.join(user_dir, f) for f in os.listdir(user_dir) if f.endswith(('.png', '.jpg', '.jpeg'))]

        if os.path.exists(user_dir):
            media_group = []
            for photo_path in photos:
                async with aiofiles.open(photo_path, 'rb') as photo_file:
                    photo_data = await photo_file.read()
                    media_group.append(InputMediaPhoto(photo_data))

        if media_group:
            media_group[0].caption = f'{name}\nИнформация: {info}'
            await bot.send_media_group(message.chat.id, media_group)

        await bot.send_message(message.chat.id, "Выберите действие:", reply_markup=markup)
        user_states[user_id]['awaiting'] = 'go_out'
        await save_user_states_async()
        logger.info(f"Профиль заказчика пользователя {user_id} отображен.")
    else:
        await bot.send_message(message.chat.id, "Профиль не найден. Пожалуйста, начните заново командой /start.")
        logger.error(f"Профиль заказчика пользователя {user_id} не найден.")
        user_states.pop(user_id, None)
        await save_user_states_async()

async def show_artist_profile(message, state):
    user_id = message.from_user.id
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute('SELECT name, style, username FROM artist WHERE tgid = ?', (str(user_id),)) as cursor:
            res = await cursor.fetchone()
    if res:
        name, style, username = res
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        edit_btn = types.KeyboardButton("Изменить профиль")
        delete_btn = types.KeyboardButton("Удалить профиль")
        delit = types.KeyboardButton("На главную")
        markup.row(edit_btn, delete_btn)
        markup.row(delit)
        user_dir = os.path.join(PHOTOS_DIR, str(user_id))
        media_group = []

        photos = [os.path.join(user_dir, f) for f in os.listdir(user_dir) if f.endswith(('.png', '.jpg', '.jpeg'))]

        if os.path.exists(user_dir):
            media_group = []
            for photo_path in photos:
                async with aiofiles.open(photo_path, 'rb') as photo_file:
                    photo_data = await photo_file.read()
                    media_group.append(InputMediaPhoto(photo_data))

        if media_group:
            media_group[0].caption = f'{name}\nОписание: {style}'
            await bot.send_media_group(message.chat.id, media_group)
        else:
            await bot.send_message(message.chat.id, f'{name}\nОписание: {style}', reply_markup=markup)

        # Получаем количество лайков и дизлайков
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute('SELECT likes, dislikes FROM ratings WHERE artist_id = ?', (str(user_id),)) as cursor:
                rating = await cursor.fetchone()

        likes = rating[0] if rating else 0
        dislikes = rating[1] if rating else 0

        await bot.send_message(message.chat.id, f'Лайки: {likes}\nДизлайки: {dislikes}', reply_markup=markup)

        await bot.send_message(message.chat.id, "Выберите действие:", reply_markup=markup)
        user_states[user_id]['awaiting'] = 'go_out'
        await save_user_states_async()
        logger.info(f"Профиль художника пользователя {user_id} отображен.")
    else:
        await bot.send_message(message.chat.id, "Профиль не найден. Пожалуйста, начните заново командой /start.")
        logger.error(f"Профиль художника пользователя {user_id} не найден.")
        user_states.pop(user_id, None)
        await save_user_states_async()

async def search_artists(message):
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute('SELECT COUNT(*) FROM artist') as cursor:
            res = await cursor.fetchone()
            count = res[0] if res else 0

    if count == 0:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        delit = types.KeyboardButton("На главную")
        markup.add(delit)
        await bot.send_message(message.chat.id, "Пока что не найдено ни одного художника.", reply_markup=markup)
        logger.info(f"Пользователь {message.from_user.id} запросил поиск художников. Найдено: {count}")
    else:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        delit = types.KeyboardButton("На главную")
        markup.add(delit)
        await bot.send_message(message.chat.id, f"Пока что найдено {count} художников.",
                               reply_markup=markup)
        logger.info(f"Пользователь {message.from_user.id} запросил поиск художников. Найдено: {count}")

        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute('SELECT tgid, name, style, username FROM artist') as cursor:
                artists = await cursor.fetchall()

        for artist in artists:
            artist_id, name, style, username = artist
            user_dir = os.path.join(PHOTOS_DIR, str(artist_id))
            media_group = []

            photos = [os.path.join(user_dir, f) for f in os.listdir(user_dir) if f.endswith(('.png', '.jpg', '.jpeg'))]

            if os.path.exists(user_dir):
                media_group = []
                for photo_path in photos:
                    async with aiofiles.open(photo_path, 'rb') as photo_file:
                        photo_data = await photo_file.read()
                        media_group.append(InputMediaPhoto(photo_data))

            if media_group:
                media_group[0].caption = f'<a href="https://t.me/{username}">{name}</a>\nОписание: {style}'
                media_group[0].parse_mode = 'HTML'
                await bot.send_media_group(message.chat.id, media_group)
            else:
                await bot.send_message(message.chat.id, f'<a href="https://t.me/{username}">{name}</a>\nОписание: {style}', parse_mode='HTML')

            # Получаем количество лайков и дизлайков
            async with aiosqlite.connect(DB_PATH) as db:
                async with db.execute('SELECT likes, dislikes FROM ratings WHERE artist_id = ?', (artist_id,)) as cursor:
                    rating = await cursor.fetchone()

            likes = rating[0] if rating else 0
            dislikes = rating[1] if rating else 0

            # Проверяем, добавлен ли художник в избранное
            async with aiosqlite.connect(DB_PATH) as db:
                async with db.execute('SELECT 1 FROM favorites WHERE customer_id = ? AND artist_id = ?', (str(message.from_user.id), artist_id)) as cursor:
                    is_favorite = await cursor.fetchone()

            like_button = types.InlineKeyboardButton(text=f"👍 {likes}", callback_data=f"like_{artist_id}")
            dislike_button = types.InlineKeyboardButton(text=f"👎 {dislikes}", callback_data=f"dislike_{artist_id}")
            favorite_button = types.InlineKeyboardButton(text="Добавить в избранное" if not is_favorite else "Удалить из избранного", callback_data=f"favorite_{artist_id}")
            keyboard = [
                [like_button, dislike_button],  # Первая строка с двумя кнопками
                [favorite_button]  # Вторая строка с одной кнопкой
            ]
            reply_markup = types.InlineKeyboardMarkup(keyboard)
            await bot.send_message(message.chat.id, f"Оцените художника:", reply_markup=reply_markup)
            await asyncio.sleep(1)  # Уменьшили задержку для лучшей производительности
            logger.info(f"Отправлена информация о художнике {name} (ID: {artist_id}).")

@bot.callback_query_handler(func=lambda call: call.data.startswith("like_") or call.data.startswith("dislike_") or call.data.startswith("favorite_"))
async def handle_rating_callback(call):
    action, artist_id = call.data.split("_")
    user_id = call.from_user.id

    async with aiosqlite.connect(DB_PATH) as db:
        if action.startswith("favorite"):
            # Проверяем, добавлен ли уже художник в избранное
            async with db.execute('SELECT 1 FROM favorites WHERE customer_id = ? AND artist_id = ?', (str(user_id), artist_id)) as cursor:
                exists = await cursor.fetchone()

            if exists:
                await db.execute('DELETE FROM favorites WHERE customer_id = ? AND artist_id = ?', (str(user_id), artist_id))
                await bot.answer_callback_query(call.id, "Художник удален из избранного.")
                logger.info(f"Пользователь {user_id} удалил художника {artist_id} из избранного.")
            else:
                await db.execute('INSERT INTO favorites (customer_id, artist_id) VALUES (?, ?)', (str(user_id), artist_id))
                await bot.answer_callback_query(call.id, "Художник добавлен в избранное.")
                logger.info(f"Пользователь {user_id} добавил художника {artist_id} в избранное.")

            await db.commit()

            # Обновляем кнопку избранного
            async with db.execute('SELECT 1 FROM favorites WHERE customer_id = ? AND artist_id = ?', (str(user_id), artist_id)) as cursor:
                is_favorite = await cursor.fetchone()

                # Получаем количество лайков и дизлайков
            async with aiosqlite.connect(DB_PATH) as db:
                async with db.execute('SELECT likes, dislikes FROM ratings WHERE artist_id = ?', (artist_id,)) as cursor:
                    rating = await cursor.fetchone()
                likes = rating[0] if rating else 0
                dislikes = rating[1] if rating else 0

            like_button = types.InlineKeyboardButton(text=f"👍 {likes}", callback_data=f"like_{artist_id}")
            dislike_button = types.InlineKeyboardButton(text=f"👎 {dislikes}", callback_data=f"dislike_{artist_id}")
            favorite_button = types.InlineKeyboardButton(text="Добавить в избранное" if not is_favorite else "Удалить из избранного", callback_data=f"favorite_{artist_id}")
            keyboard = [
                [like_button, dislike_button],  # Первая строка с двумя кнопками
                [favorite_button]  # Вторая строка с одной кнопкой
            ]
            reply_markup = types.InlineKeyboardMarkup(keyboard)
            await bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=reply_markup)

        else:
            # Проверяем, ставил ли уже пользователь лайк или дизлайк этому художнику
            async with db.execute('SELECT 1 FROM user_ratings WHERE customer_id = ? AND artist_id = ?', (str(user_id), artist_id)) as cursor:
                exists = await cursor.fetchone()

            if exists:
                await bot.answer_callback_query(call.id, "Вы уже оставили оценку для этого художника.")
                logger.info(f"Пользователь {user_id} уже оставил оценку для художника {artist_id}.")
                return

            async with db.execute('SELECT likes, dislikes FROM ratings WHERE artist_id = ?', (artist_id,)) as cursor:
                rating = await cursor.fetchone()

            likes = rating[0] if rating else 0
            dislikes = rating[1] if rating else 0

            if action == "like":
                likes += 1
            elif action == "dislike":
                dislikes += 1

            if rating:
                await db.execute('UPDATE ratings SET likes = ?, dislikes = ? WHERE artist_id = ?', (likes, dislikes, artist_id))
            else:
                await db.execute('INSERT INTO ratings (artist_id, likes, dislikes) VALUES (?, ?, ?)', (artist_id, likes, dislikes))

            await db.execute('INSERT INTO user_ratings (customer_id, artist_id, action) VALUES (?, ?, ?)', (str(user_id), artist_id, action))
            await db.commit()

            # Проверяем, добавлен ли уже художник в избранное
            async with db.execute('SELECT 1 FROM favorites WHERE customer_id = ? AND artist_id = ?',
                                  (str(user_id), artist_id)) as cursor:
                exists = await cursor.fetchone()

            if exists:
                await db.execute('DELETE FROM favorites WHERE customer_id = ? AND artist_id = ?',
                                 (str(user_id), artist_id))
                await bot.answer_callback_query(call.id, "Художник удален из избранного.")
                logger.info(f"Пользователь {user_id} удалил художника {artist_id} из избранного.")
            else:
                await db.execute('INSERT INTO favorites (customer_id, artist_id) VALUES (?, ?)',
                                 (str(user_id), artist_id))
                await bot.answer_callback_query(call.id, "Художник добавлен в избранное.")
                logger.info(f"Пользователь {user_id} добавил художника {artist_id} в избранное.")

            await db.commit()

            # Обновляем кнопку избранного
            async with db.execute('SELECT 1 FROM favorites WHERE customer_id = ? AND artist_id = ?',
                                  (str(user_id), artist_id)) as cursor:
                is_favorite = await cursor.fetchone()

            like_button = types.InlineKeyboardButton(text=f"👍 {likes}", callback_data=f"like_{artist_id}")
            dislike_button = types.InlineKeyboardButton(text=f"👎 {dislikes}", callback_data=f"dislike_{artist_id}")
            favorite_button = types.InlineKeyboardButton(text="Добавить в избранное" if not is_favorite else "Удалить из избранного", callback_data=f"favorite_{artist_id}")
            keyboard = [
                [like_button, dislike_button],  # Первая строка с двумя кнопками
                [favorite_button]  # Вторая строка с одной кнопкой
            ]
            reply_markup = types.InlineKeyboardMarkup(keyboard)
            await bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=reply_markup)

            logger.info(f"Пользователь {user_id} {action} художника {artist_id}.")

@bot.message_handler(content_types=['photo'])
async def handle_photo(message):
    user_id = message.from_user.id
    state = user_states.get(user_id, {})
    role = state.get('role')
    awaiting = state.get('awaiting')

    logger.info(f"Пользователь {user_id} отправил фотографию. Состояние: {state}")

    if role == 'customer' and awaiting == 'customer_photo':
        await handle_customer_messages(message, state)
    elif role == 'artist' and awaiting == 'artist_photo':
        await handle_artist_messages(message, state)
    else:
        await save_photos(message)
        await bot.send_message(message.chat.id, "Фотографии сохранены.")
        logger.info(f"Фотографии пользователя {user_id} сохранены.")

@bot.message_handler(content_types=['text', 'document'])
async def handle_text_and_documents(message):
    user_id = message.from_user.id
    state = user_states.get(user_id, {})

    logger.info(
        f"Пользователь {user_id} отправил сообщение или документ: {message.text if message.text else 'Документ'}. Состояние: {state}")

    # Обработка остальных сообщений
    await handle_all_messages(message)

@bot.message_handler(commands=['help'])
async def send_help(message):
    await bot.send_message(message.chat.id, "Доступные команды:\n/start - Начать\n/help - Помощь")
    logger.info(f"Пользователь {message.from_user.id} запросил помощь.")

async def save_photos(message):
    user_id = message.from_user.id
    user_dir = os.path.join(PHOTOS_DIR, str(user_id))

    # Создаем директорию, если она не существует
    if not os.path.exists(user_dir):
        os.makedirs(user_dir)

    file_info = await bot.get_file(message.photo[-1].file_id)
    file_path = file_info.file_path

    # Скачиваем фото
    file = await bot.download_file(file_path)

    # Находим максимальный номер существующего файла
    existing_files = [f for f in os.listdir(user_dir) if f.endswith(('.png', '.jpg', '.jpeg'))]
    existing_numbers = [int(f.split('.')[0]) for f in existing_files if f.split('.')[0].isdigit()]
    next_number = max(existing_numbers, default=0) + 1

    # Сохраняем фото в папку пользователя с новым именем
    file_name = os.path.join(user_dir, f"{next_number}.jpg")
    async with aiofiles.open(file_name, 'wb') as f:
        await f.write(file)

    # Сохраняем фото в состояние пользователя
    if 'photos' not in user_states[user_id]:
        user_states[user_id]['photos'] = []
    user_states[user_id]['photos'].append({'file_id': message.photo[-1].file_id, 'file_data': file})
    await save_user_states_async()

async def delete_customer_profile(message, state):
    user_id = message.from_user.id
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute('DELETE FROM customer WHERE tgid = ?', (str(user_id),))
        await db.commit()
        logger.info(f"Профиль заказчика пользователя {user_id} удален из базы данных.")

    # Удаляем фотографии пользователя
    user_dir = os.path.join(PHOTOS_DIR, str(user_id))
    if os.path.exists(user_dir):
        for file_name in os.listdir(user_dir):
            file_path = os.path.join(user_dir, file_name)
            os.remove(file_path)
        os.rmdir(user_dir)
        logger.info(f"Фотографии пользователя {user_id} удалены.")

    await bot.send_message(message.chat.id, "Ваш профиль удален.", reply_markup=types.ReplyKeyboardRemove())
    user_states.pop(user_id, None)
    await save_user_states_async()
    await send_welcome(message)

async def delete_artist_profile(message, state):
    user_id = message.from_user.id
    async with aiosqlite.connect(DB_PATH) as db:
        # Удаление из таблицы customer
        await db.execute('DELETE FROM customer WHERE tgid = ?', (str(user_id),))
        # Удаление из таблицы artist
        await db.execute('DELETE FROM artist WHERE tgid = ?', (str(user_id),))
        await db.commit()
        logger.info(f"Профиль художника пользователя {user_id} удален из базы данных.")

    # Удаляем фотографии пользователя
    user_dir = os.path.join(PHOTOS_DIR, str(user_id))
    if os.path.exists(user_dir):
        for file_name in os.listdir(user_dir):
            file_path = os.path.join(user_dir, file_name)
            os.remove(file_path)
        os.rmdir(user_dir)
        logger.info(f"Фотографии пользователя {user_id} удалены.")

    await bot.send_message(message.chat.id, "Ваш профиль удален.", reply_markup=types.ReplyKeyboardRemove())
    user_states.pop(user_id, None)
    await save_user_states_async()
    await send_welcome(message)

async def show_favorites(message):
    user_id = message.from_user.id
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute('SELECT artist.tgid, artist.name, artist.style, artist.username FROM favorites JOIN artist ON favorites.artist_id = artist.tgid WHERE favorites.customer_id = ?', (str(user_id),)) as cursor:
            favorites = await cursor.fetchall()

    if not favorites:
        await bot.send_message(message.chat.id, "В вашем избранном пока нет художников.")
        logger.info(f"Пользователь {user_id} запросил избранное. Найдено: 0 художников.")
    else:
        for artist in favorites:
            artist_id, name, style, username = artist
            user_dir = os.path.join(PHOTOS_DIR, str(artist_id))
            media_group = []

            photos = [os.path.join(user_dir, f) for f in os.listdir(user_dir) if f.endswith(('.png', '.jpg', '.jpeg'))]

            if os.path.exists(user_dir):
                media_group = []
                for photo_path in photos:
                    async with aiofiles.open(photo_path, 'rb') as photo_file:
                        photo_data = await photo_file.read()
                        media_group.append(InputMediaPhoto(photo_data))

            if media_group:
                media_group[0].caption = f'<a href="https://t.me/{username}">{name}</a>\nОписание: {style}'
                media_group[0].parse_mode = 'HTML'
                await bot.send_media_group(message.chat.id, media_group)
            else:
                await bot.send_message(message.chat.id, f'<a href="https://t.me/{username}">{name}</a>\nОписание: {style}', parse_mode='HTML')

            # Получаем количество лайков и дизлайков
            async with aiosqlite.connect(DB_PATH) as db:
                async with db.execute('SELECT likes, dislikes FROM ratings WHERE artist_id = ?', (artist_id,)) as cursor:
                    rating = await cursor.fetchone()

            likes = rating[0] if rating else 0
            dislikes = rating[1] if rating else 0

            like_button = types.InlineKeyboardButton(text=f"👍 {likes}", callback_data=f"like_{artist_id}")
            dislike_button = types.InlineKeyboardButton(text=f"👎 {dislikes}", callback_data=f"dislike_{artist_id}")
            favorite_button = types.InlineKeyboardButton(text="Удалить из избранного", callback_data=f"favorite_{artist_id}")
            keyboard = [
                [like_button, dislike_button],  # Первая строка с двумя кнопками
                [favorite_button]  # Вторая строка с одной кнопкой
            ]
            reply_markup = types.InlineKeyboardMarkup(keyboard)
            await bot.send_message(message.chat.id, f"Оцените художника:", reply_markup=reply_markup)
            await asyncio.sleep(1)  # Уменьшили задержку для лучшей производительности
            logger.info(f"Отправлена информация о художнике {name} (ID: {artist_id}).")

# Запуск бота
async def main():
    await bot.infinity_polling()

if __name__ == '__main__':
    asyncio.run(main())
